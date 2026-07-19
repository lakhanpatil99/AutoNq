"""
excel_backend.py
AutoNQ AI – Excel Data Engine
Single source of truth: data/audit_master_data.xlsx
Thread-safe read/write, Streamlit-cached, zero mock data.
"""

import os
import threading
import uuid
import base64
import logging
from datetime import datetime, date

import pandas as pd
from openpyxl import load_workbook
import streamlit as st

logger = logging.getLogger(__name__)

# ─── CONFIGURATION ────────────────────────────────────────────────────────────
EXCEL_PATH = os.path.join("data", "audit_master_data.xlsx")
_WRITE_LOCK = threading.Lock()

# Sheet name constants (must match actual workbook sheet names exactly)
SHEET_AUDIT_ENTRIES    = "audit_entries"
SHEET_LINE_MASTER      = "line_master"
SHEET_STATION_MASTER   = "station_master"
SHEET_CHECKLIST_MASTER = "checklist_master"
SHEET_SEVERITY_MASTER  = "severity_master"
SHEET_CATEGORY_MASTER  = "category_master"
SHEET_DAILY_SUMMARY    = "daily_summary"
SHEET_WEEKLY_SUMMARY   = "weekly_summary"
SHEET_MONTHLY_SUMMARY  = "monthly_summary"
SHEET_REPEATABILITY    = "repeatability_tracker"
SHEET_PROCESS_AUDIT    = "process_audit"
SHEET_FOLLOWUP         = "followup_tracker"
SHEET_EXTERNAL         = "external_tracker"
SHEET_AI_MAIL          = "ai_mail_summaries"

# Canonical audit_entries column order (matches the workbook exactly)
AUDIT_COLUMNS = [
    "audit_id", "audit_date", "audit_time", "line", "area",
    "station_no", "station_name", "checkpoint", "expected_result",
    "actual_result", "remarks", "severity", "category",
    "auditor_name", "flm_name", "shift", "status", "created_at",
    "image_base64",
]

# ── Centralized column normalization engine ──────────────────────────────────
# Maps any known Excel header variant → canonical column name.
# Used by load_audit_entries() and can be called by external modules.
_HEADER_RENAME_MAP = {
    "actual_result_observation": "actual_result",
    "actual_result___observation": "actual_result",
    "line_name": "line",
    "production_line": "line",
    "station": "station_name",
    "date": "audit_date",
    "image_path": "image_base64",
}

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize a DataFrame's columns to the canonical audit schema.
    Handles spaces, slashes, dashes, casing, and known aliases.
    This is the SINGLE centralized normalization function for the entire project.
    """
    if len(df.columns) == 0:
        return df
    # Step 1: snake_case normalization
    df.columns = [
        str(c).strip().lower()
        .replace(" / ", "_")       # "actual result / observation" → "actual result_observation"
        .replace("/", "_")
        .replace(" ", "_")
        .replace("-", "_")
        for c in df.columns
    ]
    # Step 2: Rename known variants to canonical names
    df.rename(columns=_HEADER_RENAME_MAP, inplace=True)
    return df


def inject_aliases(df: pd.DataFrame) -> pd.DataFrame:
    """Inject backward-compatible alias columns used by AI agents and app.py.
    Must be called AFTER normalize_columns().
    """
    if len(df.columns) == 0:
        return df
    # observation_text ← actual_result (the finding, not the question)
    if "observation_text" not in df.columns:
        for candidate in ["actual_result", "actual_result_observation", "actual_result___observation"]:
            if candidate in df.columns:
                df["observation_text"] = df[candidate]
                break
        else:
            if "checkpoint" in df.columns:
                df["observation_text"] = df["checkpoint"]
    # ai_principle ← category
    if "ai_principle" not in df.columns and "category" in df.columns:
        df["ai_principle"] = df["category"]
    # supervisor ← flm_name
    if "supervisor" not in df.columns and "flm_name" in df.columns:
        df["supervisor"] = df["flm_name"]
    # station alias ← station_name
    if "station" not in df.columns and "station_name" in df.columns:
        df["station"] = df["station_name"]
    # date alias ← audit_date (backward compat)
    if "date" not in df.columns and "audit_date" in df.columns:
        df["date"] = df["audit_date"]

    # Strip embedded [Auditor: ...] metadata from observation_text (legacy cleanup)
    if "observation_text" in df.columns:
        df["observation_text"] = (
            df["observation_text"].astype(str)
            .str.replace(r'\s*\[Auditor:.*?\]', '', regex=True)
            .str.strip()
        )

    return df


# ══════════════════════════════════════════════════════════════════════════════
#  LOW-LEVEL READ / WRITE
# ══════════════════════════════════════════════════════════════════════════════

def load_sheet(sheet_name: str) -> pd.DataFrame:
    """
    Read one sheet from the workbook.
    Row 1 = section-header banner  →  skipped.
    Row 2 = column headers         →  used as DataFrame columns.
    Rows 3+ = data.
    """
    if not os.path.isfile(EXCEL_PATH):
        logger.error("Excel file not found: %s", EXCEL_PATH)
        return pd.DataFrame()
    try:
        df = pd.read_excel(
            EXCEL_PATH,
            sheet_name=sheet_name,
            header=1,          # 0-indexed: second physical row
            engine="openpyxl",
        )
        df = df.dropna(how="all").reset_index(drop=True)
        # Strip whitespace from string column names
        df.columns = [str(c).strip() for c in df.columns]
        return df
    except Exception as exc:
        logger.warning("load_sheet(%s) failed: %s", sheet_name, exc)
        return pd.DataFrame()


def _safe_val(val):
    """Convert pandas/numpy types to plain Python for openpyxl."""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(val, pd.Timestamp):
        return val.to_pydatetime()
    if hasattr(val, "item"):          # numpy scalar
        return val.item()
    return val


def save_sheet(sheet_name: str, df: pd.DataFrame) -> bool:
    """
    Overwrite a sheet's data rows while preserving rows 1-2
    (section header + column header banner).
    Returns True on success.
    """
    with _WRITE_LOCK:
        try:
            wb = load_workbook(EXCEL_PATH)
            if sheet_name not in wb.sheetnames:
                logger.error("Sheet not found in workbook: %s", sheet_name)
                return False
            ws = wb[sheet_name]

            # Wipe existing data rows (keep rows 1 and 2)
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

            # Write new rows starting at row 3
            for r_offset, row_data in enumerate(df.itertuples(index=False)):
                for c_idx, val in enumerate(row_data, start=1):
                    ws.cell(row=3 + r_offset, column=c_idx, value=_safe_val(val))

            wb.save(EXCEL_PATH)
            return True
        except Exception as exc:
            logger.error("save_sheet(%s) failed: %s", sheet_name, exc)
            return False


def _append_row_to_sheet(sheet_name: str, values: list) -> bool:
    """Append a single row to a sheet (does NOT acquire _WRITE_LOCK — caller must hold it)."""
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb[sheet_name]
        next_row = ws.max_row + 1
        for c_idx, val in enumerate(values, start=1):
            ws.cell(row=next_row, column=c_idx, value=_safe_val(val))
        wb.save(EXCEL_PATH)
        return True
    except Exception as exc:
        logger.error("_append_row_to_sheet(%s) failed: %s", sheet_name, exc)
        return False


# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT-CACHED READERS
#  TTL values chosen per data volatility:
#    audit_entries → 30s  (written frequently)
#    master tables → 300s (rarely change)
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=30, show_spinner=False)
def load_audit_entries() -> pd.DataFrame:
    df = load_sheet(SHEET_AUDIT_ENTRIES)
    if df.empty:
        logger.warning("load_audit_entries: sheet is empty — returning schema-only DataFrame")
        return pd.DataFrame(columns=AUDIT_COLUMNS)

    # ── Step 1: Centralized column normalization ─────────────────────────────
    df = normalize_columns(df)

    # ── Step 2: Date parsing ─────────────────────────────────────────────────
    for col in ("created_at",):
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    if "audit_date" in df.columns:
        df["audit_date"] = pd.to_datetime(df["audit_date"], errors="coerce")
        # FIX: Fallback NaT dates to today to prevent silent row drops in date filters
        df["audit_date"] = df["audit_date"].fillna(pd.Timestamp.today())

    # ── Step 3: Inject backward-compatible aliases ───────────────────────────
    df = inject_aliases(df)

    # ── Validation logger ────────────────────────────────────────────────────
    logger.info(
        "load_audit_entries: %d rows, %d cols | has actual_result=%s, observation_text=%s, line=%s",
        len(df), len(df.columns),
        "actual_result" in df.columns,
        "observation_text" in df.columns,
        "line" in df.columns,
    )
    return df


@st.cache_data(ttl=300, show_spinner=False)
def _load_line_master_raw() -> pd.DataFrame:
    return load_sheet(SHEET_LINE_MASTER)


@st.cache_data(ttl=300, show_spinner=False)
def _load_station_master_raw() -> pd.DataFrame:
    return load_sheet(SHEET_STATION_MASTER)


@st.cache_data(ttl=300, show_spinner=False)
def _load_checklist_master_raw() -> pd.DataFrame:
    return load_sheet(SHEET_CHECKLIST_MASTER)


@st.cache_data(ttl=600, show_spinner=False)
def _load_severity_master_raw() -> pd.DataFrame:
    return load_sheet(SHEET_SEVERITY_MASTER)


@st.cache_data(ttl=600, show_spinner=False)
def _load_category_master_raw() -> pd.DataFrame:
    return load_sheet(SHEET_CATEGORY_MASTER)


def refresh_master_data():
    """Bust all caches to force a fresh read on the next call."""
    load_audit_entries.clear()
    _load_line_master_raw.clear()
    _load_station_master_raw.clear()
    _load_checklist_master_raw.clear()
    _load_severity_master_raw.clear()
    _load_category_master_raw.clear()


# ══════════════════════════════════════════════════════════════════════════════
#  DYNAMIC DROPDOWN HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _col(df: pd.DataFrame, *keywords) -> str | None:
    """Find the first column whose name contains any of the given keywords (case-insensitive)."""
    for col in df.columns:
        cl = col.lower()
        if any(k in cl for k in keywords):
            return col
    return None


def get_line_options() -> list[str]:
    """Return active line names from line_master, sorted."""
    df = _load_line_master_raw()
    if df.empty:
        return []
    name_col   = _col(df, "name")
    status_col = _col(df, "status")
    if name_col is None:
        return []
    if status_col:
        df = df[df[status_col].astype(str).str.strip().str.lower() == "active"]
    return sorted(df[name_col].dropna().astype(str).str.strip().tolist())


def get_stations_for_line(line: str) -> list[str]:
    """Return station names for the given line from station_master."""
    df = _load_station_master_raw()
    if df.empty:
        return []
    line_col    = _col(df, "line")
    station_col = _col(df, "name")
    if not line_col or not station_col:
        return []
    mask = df[line_col].astype(str).str.strip() == line.strip()
    return sorted(df.loc[mask, station_col].dropna().astype(str).str.strip().tolist())


def get_station_no(line: str, station_name: str) -> str:
    """Return the station number string for a given line + station name."""
    df = _load_station_master_raw()
    if df.empty:
        return ""
    line_col    = _col(df, "line")
    name_col    = _col(df, "name")
    no_col      = _col(df, "no", "number")
    if not all([line_col, name_col, no_col]):
        return ""
    mask = (
        (df[line_col].astype(str).str.strip() == line.strip()) &
        (df[name_col].astype(str).str.strip() == station_name.strip())
    )
    rows = df.loc[mask, no_col]
    return str(rows.iloc[0]).strip() if not rows.empty else ""


def get_checklist_for_station(line: str, station_name: str) -> pd.DataFrame:
    """
    Return active checklist rows for the given line + station.
    Columns of interest: Checklist ID, Checkpoint, Expected Result, Category, Severity.
    """
    df = _load_checklist_master_raw()
    if df.empty:
        return pd.DataFrame()
    line_col    = _col(df, "line")
    station_col = _col(df, "station name", "stationname")
    active_col  = _col(df, "active")
    if not line_col or not station_col:
        return df
    mask = (
        (df[line_col].astype(str).str.strip() == line.strip()) &
        (df[station_col].astype(str).str.strip() == station_name.strip())
    )
    filtered = df[mask].copy()
    if active_col:
        filtered = filtered[filtered[active_col].astype(str).str.strip().str.upper() == "Y"]
    return filtered.reset_index(drop=True)


def get_severity_options() -> list[str]:
    df = _load_severity_master_raw()
    if df.empty:
        return []  # No fallback — UI must handle empty list
    col = df.columns[0]
    return df[col].dropna().astype(str).str.strip().tolist()


def get_category_options() -> list[str]:
    df = _load_category_master_raw()
    if df.empty:
        return []  # No fallback — UI must handle empty list
    col = df.columns[0]
    return df[col].dropna().astype(str).str.strip().tolist()


# ══════════════════════════════════════════════════════════════════════════════
#  AUDIT ENTRY APPEND
# ══════════════════════════════════════════════════════════════════════════════

def _generate_audit_id() -> str:
    ts  = datetime.now().strftime("%Y%m%d%H%M%S")
    uid = str(uuid.uuid4())[:6].upper()
    return f"AUD-{ts}-{uid}"






def add_audit_entry(entry_data: dict) -> bool:
    """
    Public interface used by app.py to persist a single audit observation.

    Accepts the field schema produced by the Audit Entry page:
        line             – production line identifier
        station          – station name (free text or from master)
        station_no       – station number (from master lookup)
        area             – production area
        supervisor       – FLM / supervisor name
        auditor          – auditor name (saved to auditor_name column)
        observation_text – clean observation description (no embedded metadata)
        ai_principle     – principle detected by detect_principle()
        severity         – user-selected severity level
        category         – user-selected category
        remarks          – optional remarks
        audit_date       – audit_date string or date object
        shift            – Shift 1 / 2 / 3
        image            – base64-encoded image string OR raw bytes OR None

    Returns True on success, False on any failure.
    Thread-safe via _WRITE_LOCK.
    """
    if not entry_data.get("line") or not entry_data.get("observation_text"):
        logger.warning("add_audit_entry: missing required fields (line / observation_text).")
        return False

    try:
        audit_id = _generate_audit_id()
        now      = datetime.now()

        # ── Normalise image to base64 string ──────────────────────────────────
        image_raw = entry_data.get(
            "image_base64",
            entry_data.get("image", "")
        )
        if isinstance(image_raw, (bytes, bytearray)):
            # raw bytes → encode
            image_b64 = base64.b64encode(image_raw).decode("utf-8")
        elif isinstance(image_raw, str):
            # already a base64 string (passed from app.py after manual encode)
            image_b64 = image_raw
        else:
            image_b64 = ""

        # ── Build row matching AUDIT_COLUMNS order ────────────────────────────
        row_values = [
            audit_id,                                   # audit_id
            entry_data.get("audit_date", date.today()),       # audit_date
            now.strftime("%H:%M"),                      # audit_time
            entry_data.get("line", ""),                 # line
            entry_data.get("area", ""),                 # area
            entry_data.get("station_no", ""),           # station_no
            entry_data.get("station", ""),              # station_name
            entry_data.get("ai_principle", ""),         # checkpoint  (principle maps here)
            "",                                         # expected_result
            entry_data.get("observation_text", ""),     # actual_result
            entry_data.get("remarks", ""),              # remarks
            entry_data.get("severity", ""),             # severity
            entry_data.get("category", ""),             # category
            entry_data.get("auditor", ""),              # auditor_name
            entry_data.get("supervisor", ""),           # flm_name
            entry_data.get("shift", ""),                # shift
            "Open",                                     # status
            now,                                        # created_at
            image_b64,                                  # image_base64
        ]

        with _WRITE_LOCK:
            # ── Ensure workbook and sheet exist ───────────────────────────────
            if not os.path.isfile(EXCEL_PATH):
                logger.error("add_audit_entry: Excel file not found at %s", EXCEL_PATH)
                return False

            wb = load_workbook(EXCEL_PATH)

            if SHEET_AUDIT_ENTRIES not in wb.sheetnames:
                # Auto-create the sheet with a header row
                ws = wb.create_sheet(SHEET_AUDIT_ENTRIES)
                # Row 1 – section banner
                ws.cell(row=1, column=1, value="AutoNQ AI – Audit Entries")
                # Row 2 – column headers
                for c_idx, col_name in enumerate(AUDIT_COLUMNS, start=1):
                    ws.cell(row=2, column=c_idx, value=col_name)
                wb.save(EXCEL_PATH)
                logger.info("add_audit_entry: created missing sheet '%s'", SHEET_AUDIT_ENTRIES)

            # Re-load wb after potential save above
            wb = load_workbook(EXCEL_PATH)
            ws = wb[SHEET_AUDIT_ENTRIES]

            next_row = ws.max_row + 1
            for c_idx, val in enumerate(row_values, start=1):
                ws.cell(row=next_row, column=c_idx, value=_safe_val(val))

            wb.save(EXCEL_PATH)

        # Bust the cached read so the UI reflects new data immediately
        load_audit_entries.clear()
        logger.info("add_audit_entry: saved audit_id=%s to row %s", audit_id, next_row)
        return True

    except Exception as exc:
        logger.error("add_audit_entry failed: %s", exc, exc_info=True)
        return False


# ══════════════════════════════════════════════════════════════════════════════
#  DATA ADAPTER  →  AI agent functions expect specific column names
# ══════════════════════════════════════════════════════════════════════════════

def get_audit_df_for_ai(days: int | None = None) -> pd.DataFrame:
    """
    Return audit_entries mapped to the column schema expected by AI agents:
    line, station, supervisor, observation_text, ai_principle, audit_date, shift, severity
    """
    df = load_audit_entries()
    if df.empty:
        return pd.DataFrame(columns=["line", "station", "supervisor",
                                     "observation_text", "ai_principle",
                                     "audit_date", "shift", "severity"])

    mapped = pd.DataFrame()
    mapped["line"]             = df.get("line",           pd.Series(dtype=str))
    mapped["station"]          = df.get("station_name",   df.get("station", pd.Series(dtype=str)))
    mapped["supervisor"]       = df.get("flm_name",       df.get("supervisor", pd.Series(dtype=str)))
    mapped["observation_text"] = df.get("observation_text", df.get("actual_result", df.get("checkpoint", pd.Series(dtype=str))))
    mapped["ai_principle"]     = df.get("category",       pd.Series(dtype=str))
    mapped["audit_date"]             = pd.to_datetime(df.get("audit_date", pd.Series(dtype=str)), errors="coerce")
    mapped["shift"]            = df.get("shift",          pd.Series(dtype=str))
    mapped["severity"]         = df.get("severity",       pd.Series(dtype=str))

    mapped = mapped.dropna(subset=["line"]).reset_index(drop=True)

    if days is not None:
        cutoff = pd.Timestamp.now() - pd.Timedelta(days=days)
        mapped = mapped[mapped["audit_date"] >= cutoff]

    return mapped


def get_current_and_previous(window_days: int = 30):
    """
    Split audit data into current window and prior window.
    Returns (current_df, previous_df, full_df).
    """
    full = get_audit_df_for_ai()
    if full.empty:
        empty = pd.DataFrame(columns=full.columns)
        return empty, empty, empty
    cutoff   = pd.Timestamp.now() - pd.Timedelta(days=window_days)
    current  = full[full["audit_date"] >= cutoff].copy()
    previous = full[full["audit_date"] <  cutoff].copy()
    return current, previous, full


def get_risk_summary_for_plan() -> pd.DataFrame:
    """
    Return per-line risk aggregation for classify_by_percentile:
    Columns: LINE, failure_count (from non-closed audits), risk_score
    """
    df = get_audit_df_for_ai()
    if df.empty:
        return pd.DataFrame(columns=["LINE", "failure_count", "risk_score"])
    agg = df.groupby("line").agg(
        failure_count=("observation_text", "count"),
    ).reset_index().rename(columns={"line": "LINE"})
    agg["risk_score"] = agg["failure_count"]
    return agg


def get_empty_iqis() -> tuple[pd.DataFrame, pd.DataFrame]:
    """Return empty DataFrames matching the IQIS schema (no longer mocked)."""
    iqis_df = pd.DataFrame(columns=["LINE", "failure_mode", "count"])
    risk_df = pd.DataFrame(columns=["LINE", "failure_count"])
    return iqis_df, risk_df


def get_empty_lpc() -> pd.DataFrame:
    return pd.DataFrame(columns=["LINE", "issue", "count"])


# ══════════════════════════════════════════════════════════════════════════════
#  SUMMARY ENGINES
# ══════════════════════════════════════════════════════════════════════════════

def compute_daily_summary() -> pd.DataFrame:
    """Aggregate audit_entries into a daily summary table."""
    df = load_audit_entries()
    if df.empty or "audit_date" not in df.columns:
        return pd.DataFrame()
    df["audit_date"] = pd.to_datetime(df["audit_date"], errors="coerce")
    df["audit_date"]      = df["audit_date"].dt.date
    line_col   = "line"   if "line"  in df.columns else None
    shift_col  = "shift"  if "shift" in df.columns else None
    status_col = "status" if "status" in df.columns else None
    sev_col    = "severity" if "severity" in df.columns else None
    aud_col    = "auditor_name" if "auditor_name" in df.columns else None

    grp_keys = ["audit_date"]
    if line_col:  grp_keys.append(line_col)
    if shift_col: grp_keys.append(shift_col)

    agg = df.groupby(grp_keys).agg(
        Total_Checkpoints=("audit_id", "count"),
        Pass   =(status_col or "audit_id", lambda x: (x.astype(str).str.lower() == "closed").sum() if status_col else 0),
        Fail   =(status_col or "audit_id", lambda x: (x.astype(str).str.lower().isin(["open","in progress"])).sum() if status_col else 0),
        Open_NCRs    =(status_col or "audit_id", lambda x: (x.astype(str).str.lower() == "open").sum() if status_col else 0),
        In_Progress  =(status_col or "audit_id", lambda x: (x.astype(str).str.lower() == "in progress").sum() if status_col else 0),
        Critical_Findings=(sev_col or "audit_id", lambda x: (x.astype(str).str.lower() == "critical").sum() if sev_col else 0),
        High_Findings    =(sev_col or "audit_id", lambda x: (x.astype(str).str.lower() == "high").sum() if sev_col else 0),
        Auditor=(aud_col or "audit_id", "first"),
    ).reset_index()

    agg["Pass_Rate_Pct"] = (
        agg["Pass"] / agg["Total_Checkpoints"].replace(0, 1)
    ).mul(100).round(1)

    rename = {"audit_date": "Summary Date"}
    if line_col:  rename[line_col]  = "Line"
    if shift_col: rename[shift_col] = "Shift"
    agg.rename(columns=rename, inplace=True)
    return agg


def compute_weekly_summary() -> pd.DataFrame:
    df = load_audit_entries()
    if df.empty or "audit_date" not in df.columns:
        return pd.DataFrame()
    df["audit_date"] = pd.to_datetime(df["audit_date"], errors="coerce")
    df["_week"]  = df["audit_date"].dt.isocalendar().week.astype(int)
    df["_year"]  = df["audit_date"].dt.year
    df["_wstart"]= df["audit_date"] - pd.to_timedelta(df["audit_date"].dt.dayofweek, unit="D")
    df["_wend"]  = df["_wstart"] + pd.Timedelta(days=6)
    line_col     = "line" if "line" in df.columns else None
    status_col   = "status" if "status" in df.columns else None
    sev_col      = "severity" if "severity" in df.columns else None

    grp = ["_year", "_week"]
    if line_col: grp.append(line_col)
    agg = df.groupby(grp).agg(
        Total_Audits=("audit_id", "count"),
        Pass =(status_col or "audit_id", lambda x: (x.astype(str).str.lower() == "closed").sum() if status_col else 0),
        Fail =(status_col or "audit_id", lambda x: (x.astype(str).str.lower().isin(["open","in progress"])).sum() if status_col else 0),
        Critical_NCRs=(sev_col or "audit_id", lambda x: (x.astype(str).str.lower() == "critical").sum() if sev_col else 0),
        Week_Start=("_wstart", "min"),
        Week_End  =("_wend",   "max"),
    ).reset_index()
    agg["Pass_Rate_Pct"] = (agg["Pass"] / agg["Total_Audits"].replace(0, 1)).mul(100).round(1)
    agg["Status"] = agg["Pass_Rate_Pct"].apply(
        lambda r: "On Target" if r >= 95 else ("Needs Attention" if r >= 85 else "Critical")
    )
    rename = {"_year": "Year", "_week": "Week No", "Week_Start": "Week Start", "Week_End": "Week End"}
    if line_col: rename[line_col] = "Line"
    agg.rename(columns=rename, inplace=True)
    return agg


def compute_monthly_summary() -> pd.DataFrame:
    df = load_audit_entries()
    if df.empty or "audit_date" not in df.columns:
        return pd.DataFrame()
    df["audit_date"] = pd.to_datetime(df["audit_date"], errors="coerce")
    df["_month"] = df["audit_date"].dt.strftime("%B")
    df["_year"]  = df["audit_date"].dt.year
    line_col   = "line"     if "line"     in df.columns else None
    status_col = "status"   if "status"   in df.columns else None
    sev_col    = "severity" if "severity" in df.columns else None
    remarks_col= "remarks"  if "remarks"  in df.columns else None

    grp = ["_year", "_month"]
    if line_col: grp.append(line_col)
    agg = df.groupby(grp).agg(
        Total   =("audit_id", "count"),
        Pass    =(status_col or "audit_id", lambda x: (x.astype(str).str.lower() == "closed").sum() if status_col else 0),
        Fail    =(status_col or "audit_id", lambda x: (x.astype(str).str.lower().isin(["open","in progress"])).sum() if status_col else 0),
        Critical_NCRs=(sev_col or "audit_id", lambda x: (x.astype(str).str.lower() == "critical").sum() if sev_col else 0),
        Repeat_Failures=(remarks_col or "audit_id",
            lambda x: x.astype(str).str.contains("repeat|recurring|again", case=False, na=False).sum() if remarks_col else 0),
    ).reset_index()
    agg["Pass_Rate_Pct"] = (agg["Pass"] / agg["Total"].replace(0, 1)).mul(100).round(1)
    agg["Target_Pct"]    = 95.0
    agg["Vs_Target"]     = agg["Pass_Rate_Pct"].apply(
        lambda r: "✔ On Target" if r >= 95 else "✘ Below Target"
    )
    rename = {"_year": "Year", "_month": "Month", "Total": "Total Checkpoints",
              "Critical_NCRs": "Critical NCRs", "Repeat_Failures": "Repeat Failures",
              "Pass_Rate_Pct": "Pass Rate %", "Target_Pct": "Target Pass Rate %",
              "Vs_Target": "Vs Target"}
    if line_col: rename[line_col] = "Line"
    agg.rename(columns=rename, inplace=True)
    return agg


def save_summary(summary_type: str, df: pd.DataFrame) -> bool:
    """Persist a computed summary DataFrame to the appropriate sheet."""
    mapping = {
        "daily":   SHEET_DAILY_SUMMARY,
        "weekly":  SHEET_WEEKLY_SUMMARY,
        "monthly": SHEET_MONTHLY_SUMMARY,
    }
    sheet = mapping.get(summary_type.lower())
    return save_sheet(sheet, df) if sheet else False


# ══════════════════════════════════════════════════════════════════════════════
#  REPEATABILITY ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def compute_repeatability(days: int = 30) -> pd.DataFrame:
    """
    Detect checkpoints that have failed 2+ times within the last `days` days.
    Returns a DataFrame sorted by failure count descending.
    """
    df = load_audit_entries()
    if df.empty:
        return pd.DataFrame()

    df["audit_date"] = pd.to_datetime(df.get("audit_date", pd.Series(dtype=str)), errors="coerce")
    cutoff = pd.Timestamp.now() - pd.Timedelta(days=days)
    df = df[df["audit_date"] >= cutoff].copy()
    if df.empty:
        return pd.DataFrame()

    line_col      = "line"         if "line"         in df.columns else None
    station_col   = "station_name" if "station_name" in df.columns else None
    checkpoint_col= "checkpoint"   if "checkpoint"   in df.columns else \
                    ("actual_result" if "actual_result" in df.columns else None)
    sev_col       = "severity"     if "severity"     in df.columns else None

    if not checkpoint_col:
        return pd.DataFrame()

    grp = [c for c in [line_col, station_col, checkpoint_col] if c]
    counts = df.groupby(grp).agg(
        Failure_Count=("audit_id",   "count"),
        Last_Failure =("audit_date", "max"),
        Severity     =(sev_col or checkpoint_col, "first"),
    ).reset_index()

    repeat = counts[counts["Failure_Count"] >= 2].copy()
    repeat["Recurrence_Risk"] = repeat["Failure_Count"].apply(
        lambda x: "Critical" if x >= 5 else "High" if x >= 3 else "Medium"
    )
    repeat = repeat.sort_values("Failure_Count", ascending=False).reset_index(drop=True)
    rename = {}
    if line_col:       rename[line_col]       = "Line"
    if station_col:    rename[station_col]    = "Station Name"
    if checkpoint_col: rename[checkpoint_col] = "Checkpoint"
    repeat.rename(columns=rename, inplace=True)
    return repeat


def save_repeatability(df: pd.DataFrame) -> bool:
    return save_sheet(SHEET_REPEATABILITY, df)


# ══════════════════════════════════════════════════════════════════════════════
#  FOLLOW-UP APPEND
# ══════════════════════════════════════════════════════════════════════════════

def append_followup_entry(ncr_data: dict) -> bool:
    """
    Append a follow-up / NCR row to followup_tracker.
    ncr_data keys: linked_audit_id, line, station, finding,
                   root_cause, corrective_action, responsible, due_date
    """
    ncr_id = f"NCR-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    row = [
        ncr_id,
        ncr_data.get("linked_audit_id", ""),
        datetime.now(),
        ncr_data.get("line", ""),
        ncr_data.get("station", ""),
        ncr_data.get("finding", ""),
        ncr_data.get("root_cause", ""),
        ncr_data.get("corrective_action", ""),
        ncr_data.get("responsible", ""),
        ncr_data.get("due_date", ""),
        None,   # completion_date
        None,   # verified_by
        "Open",
    ]
    with _WRITE_LOCK:
        return _append_row_to_sheet(SHEET_FOLLOWUP, row)


# ══════════════════════════════════════════════════════════════════════════════
#  AI MAIL SUMMARY LOG
# ══════════════════════════════════════════════════════════════════════════════

def log_ai_mail_summary(summary_type: str, lines_covered: str, narrative: str,
                        critical_count: int, open_ncrs: int, recipients: str) -> bool:
    """Append a generated AI mail summary record to ai_mail_summaries sheet."""
    summary_id = f"AISUM-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    row = [
        summary_id,
        datetime.now(),
        summary_type,
        lines_covered,
        narrative[:2000] if narrative else "",  # cap at 2000 chars per cell
        critical_count,
        open_ncrs,
        recipients,
        "Sent",
    ]
    with _WRITE_LOCK:
        return _append_row_to_sheet(SHEET_AI_MAIL, row)


# ══════════════════════════════════════════════════════════════════════════════
#  CONVENIENCE LOAD ALL
# ══════════════════════════════════════════════════════════════════════════════

def load_master_data() -> dict:
    """Return all sheets as a dict of DataFrames."""
    return {
        SHEET_AUDIT_ENTRIES:    load_audit_entries(),
        SHEET_LINE_MASTER:      _load_line_master_raw(),
        SHEET_STATION_MASTER:   _load_station_master_raw(),
        SHEET_CHECKLIST_MASTER: _load_checklist_master_raw(),
        SHEET_SEVERITY_MASTER:  _load_severity_master_raw(),
        SHEET_CATEGORY_MASTER:  _load_category_master_raw(),
    }