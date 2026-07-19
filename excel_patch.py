import openpyxl
from copy import copy
import json

file_path = r"d:\HACK28\data\audit_master_data.xlsx"

wb = openpyxl.load_workbook(file_path)

def append_rows_to_sheet(sheet_name, match_col_idx, match_val, new_vals_dict):
    """
    Finds rows matching `match_val` in column `match_col_idx` (1-indexed).
    Copies those rows, updating the values based on `new_vals_dict`, and appends them.
    `new_vals_dict` maps column index (1-indexed) to a list of new values.
    Returns the number of rows appended for reporting.
    """
    if sheet_name not in wb.sheetnames:
        return []
    
    ws = wb[sheet_name]
    
    # Identify matching rows (ignoring header rows 1 and 2)
    matching_rows = []
    for row in ws.iter_rows(min_row=3):
        if row[match_col_idx - 1].value == match_val:
            matching_rows.append(row)
    
    if not matching_rows:
        return []
        
    appended_reports = []
    
    for i in range(len(new_vals_dict[list(new_vals_dict.keys())[0]])):
        for original_row in matching_rows:
            new_row_values = []
            for cell in original_row:
                new_row_values.append(cell.value)
                
            # Apply overrides for the i-th duplicate
            for col_idx, vals in new_vals_dict.items():
                new_row_values[col_idx - 1] = vals[i]
                
            ws.append(new_row_values)
            appended_reports.append(new_row_values)
            
            # Copy styling (optional but safe)
            new_row_idx = ws.max_row
            for col_idx, cell in enumerate(original_row, start=1):
                new_cell = ws.cell(row=new_row_idx, column=col_idx)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
                    
    return appended_reports


# 1. line_master 
# Col 1 = Line ID, Col 2 = Line Name
line_master_report = append_rows_to_sheet(
    "line_master", 
    match_col_idx=1, # Match on Line ID column
    match_val="L002", 
    new_vals_dict={
        1: ["L008", "L009"], 
        2: ["Line 5", "Line 7"]
    }
)

# 2. station_master
# Col 1 = Line, Col 2 = Station No
station_master_report = append_rows_to_sheet(
    "station_master",
    match_col_idx=1,
    match_val="Line 2",
    new_vals_dict={
        1: ["Line 5", "Line 7"]
    }
)

# 3. checklist_master
# Col 1 = Checkpoint, Col 2 = Line, Col 3 = Station ... wait, let's check col index for checklist_master.
# In my previous run: checklist_master cols: 
# 'CHECKLIST MASTER – Bosch Audit Checkpoint Library | MASTER DATA DRIVER', 'Unnamed: 1', 'Unnamed: 2'...
# Wait, header is at row 2! 
