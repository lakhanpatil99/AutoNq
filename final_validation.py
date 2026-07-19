"""
FINAL PRODUCTION READINESS VALIDATION
AutoNQ AI — Image Ownership System
Covers: Phases 1-9 (Stress, Multi-Station, Multi-Line, Excel, Session, Regression)
"""
import sys, os, base64, glob, hashlib
os.environ["STREAMLIT_SERVER_HEADLESS"] = "true"
import openpyxl
import pandas as pd
from datetime import date

from excel_backend import (
    add_audit_entry, load_audit_entries, AUDIT_COLUMNS,
    EXCEL_PATH, SHEET_AUDIT_ENTRIES, normalize_columns, inject_aliases
)

P, F = "PASS", "FAIL"
results = []
def T(name, cond, detail=""):
    s = P if cond else F
    results.append((name, s, detail))
    mark = "+" if s == P else "X"
    print(f"  [{mark}] {name}" + (f" | {detail}" if detail else ""))

# ====================================================================
print("="*72)
print("PHASE 1: CODEBASE AUDIT — image keyword scan (production files only)")
print("="*72)
prod = ['app.py','excel_backend.py','setup_environment.py','ui_styles.py','mail_service.py','get_icons.py']
dangerous = ['_img_lookup','station_image','line_image','image_cache']
for pat in dangerous:
    for f in prod:
        if os.path.exists(f):
            content = open(f, encoding='utf-8').read()
            if pat in content:
                print(f"  ALERT: '{pat}' found in {f}")
                T(f"No '{pat}' in {f}", False)
            else:
                T(f"No '{pat}' in {f}", True)

# ====================================================================
print("\n" + "="*72)
print("PHASE 2: SCHEMA VERIFICATION")
print("="*72)
T("image_base64 in AUDIT_COLUMNS", "image_base64" in AUDIT_COLUMNS,
  f"index={AUDIT_COLUMNS.index('image_base64')}")

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb[SHEET_AUDIT_ENTRIES]
hdrs = [c.value for c in ws[2]]
T("'Image Base64' in Excel headers", "Image Base64" in hdrs,
  f"col={hdrs.index('Image Base64')+1 if 'Image Base64' in hdrs else '?'}")
wb.close()

# ====================================================================
print("\n" + "="*72)
print("PHASE 3: STRESS TEST — 5 obs, SAME station, 5 UNIQUE images")
print("="*72)

imgs = {
    "RED":    base64.b64encode(b'STRESS_RED_001').decode(),
    "BLUE":   base64.b64encode(b'STRESS_BLUE_002').decode(),
    "GREEN":  base64.b64encode(b'STRESS_GREEN_003').decode(),
    "YELLOW": base64.b64encode(b'STRESS_YELLOW_004').decode(),
    "BLACK":  base64.b64encode(b'STRESS_BLACK_005').decode(),
}
for color, b64 in imgs.items():
    print(f"  {color}: {b64}")

for color, b64 in imgs.items():
    ok = add_audit_entry({
        'line': 'STRESS_LINE', 'station': 'SAME_STATION',
        'observation_text': f'Stress obs {color}',
        'image_base64': b64, 'audit_date': '2026-05-29'
    })
    T(f"Save {color}", ok)

load_audit_entries.clear()
df = load_audit_entries()
stress = df[df['line'] == 'STRESS_LINE'].copy()
T("5 stress rows loaded", len(stress) == 5, f"got {len(stress)}")

saved_imgs = list(stress['image_base64'])
T("All 5 images unique in DataFrame", len(set(str(x) for x in saved_imgs)) == 5,
  f"unique={len(set(str(x) for x in saved_imgs))}")

for i, (color, expected) in enumerate(imgs.items()):
    actual = str(saved_imgs[i]) if i < len(saved_imgs) else ""
    T(f"Row {i} == {color}", actual == expected,
      f"expected={expected[:20]}... actual={actual[:20]}...")

# Raw Excel verification
wb2 = openpyxl.load_workbook(EXCEL_PATH)
ws2 = wb2[SHEET_AUDIT_ENTRIES]
raw_stress = []
for r in range(3, ws2.max_row + 1):
    if str(ws2.cell(r, 4).value).strip() == 'STRESS_LINE':
        raw_stress.append(ws2.cell(r, 19).value)
wb2.close()
T("Raw Excel has 5 stress images", len(raw_stress) == 5, f"got {len(raw_stress)}")
T("Raw Excel images all unique", len(set(raw_stress)) == 5)

# ====================================================================
print("\n" + "="*72)
print("PHASE 4: MULTI-STATION TEST — 3 stations, 3 images")
print("="*72)

ms_imgs = {
    "STN_ALPHA": base64.b64encode(b'MS_ALPHA_IMG').decode(),
    "STN_BETA":  base64.b64encode(b'MS_BETA_IMG').decode(),
    "STN_GAMMA": base64.b64encode(b'MS_GAMMA_IMG').decode(),
}
for stn, b64 in ms_imgs.items():
    add_audit_entry({
        'line': 'MS_LINE', 'station': stn,
        'observation_text': f'Multi-station obs at {stn}',
        'image_base64': b64, 'audit_date': '2026-05-29'
    })

load_audit_entries.clear()
df2 = load_audit_entries()
ms = df2[df2['line'] == 'MS_LINE'].copy()
T("3 multi-station rows loaded", len(ms) == 3, f"got {len(ms)}")

for _, row in ms.iterrows():
    stn = str(row.get('station', row.get('station_name', '')))
    img = str(row['image_base64'])
    expected = ms_imgs.get(stn, "")
    T(f"Station {stn} has correct image", img == expected)

# ====================================================================
print("\n" + "="*72)
print("PHASE 5: MULTI-LINE TEST — 3 lines, 3 images")
print("="*72)

ml_imgs = {
    "ML_LINE_X": base64.b64encode(b'ML_X_IMG').decode(),
    "ML_LINE_Y": base64.b64encode(b'ML_Y_IMG').decode(),
    "ML_LINE_Z": base64.b64encode(b'ML_Z_IMG').decode(),
}
for line, b64 in ml_imgs.items():
    add_audit_entry({
        'line': line, 'station': 'SHARED_STN',
        'observation_text': f'Multi-line obs on {line}',
        'image_base64': b64, 'audit_date': '2026-05-29'
    })

load_audit_entries.clear()
df3 = load_audit_entries()
for line, expected in ml_imgs.items():
    row = df3[df3['line'] == line]
    if row.empty:
        T(f"Line {line} loaded", False, "no rows")
        continue
    actual = str(row['image_base64'].iloc[0])
    T(f"Line {line} has correct image (no line inheritance)", actual == expected)

# ====================================================================
print("\n" + "="*72)
print("PHASE 6: generate_qcheck_questions() CODE PATH SIMULATION")
print("="*72)

# Replicate EXACT code from setup_environment.py:712-761
stress_records = stress.to_dict(orient="records")
filtered = [r for r in stress_records if str(r.get("line")) == "STRESS_LINE"]
filtered = sorted(filtered, key=lambda x: x.get("date", ""), reverse=True)
seen = set()
selected = []
for row in filtered:
    key = (row.get("station"), row.get("observation_text"))
    if key not in seen:
        seen.add(key)
        selected.append(row)
    if len(selected) >= 15:
        break

# Simulate AI output: 1 question per selected row
sim_questions = []
for qi, row in enumerate(selected):
    q = f"Simulated Q{qi}"
    sim_questions.append({
        "Station": row.get("station", row.get("station_name", "")),
        "Checkpoint": q,
        "Ref Photo": "", "Status": "OK", "Remark": "",
        "image_base64": row.get("image_base64", "")  # PATCHED LINE
    })

T("5 simulated Q-Check records generated", len(sim_questions) == 5,
  f"got {len(sim_questions)}")

sim_imgs = [sq["image_base64"] for sq in sim_questions]
T("All Q-Check images unique (no station collapse)", len(set(str(x) for x in sim_imgs)) == 5)

for i, (color, expected) in enumerate(imgs.items()):
    if i < len(sim_questions):
        actual = str(sim_questions[i].get("image_base64", ""))
        T(f"Q-Check Q{i} image == {color}", actual == expected)

# ====================================================================
print("\n" + "="*72)
print("PHASE 7: FRONTEND RENDER PATH VERIFICATION")
print("="*72)
# Verify the render code reads from row directly, no lookup
app_code = open('app.py', encoding='utf-8').read()
T("Render reads row.get('image_base64')",
  'row.get("image_base64")' in app_code or "row.get('image_base64')" in app_code)
T("No _img_lookup in app.py", "_img_lookup" not in app_code)
T("No station image lookup in app.py",
  "station" not in app_code.split("# PAGE: DAILY Q-CHECK")[1].split("# PAGE: PROCESS AUDIT")[0].lower()
  if "# PAGE: DAILY Q-CHECK" in app_code and "# PAGE: PROCESS AUDIT" in app_code
  else False)

# ====================================================================
print("\n" + "="*72)
print("PHASE 8: NOK SUBMISSION — NO LEAKAGE, NO NAMEERROR")
print("="*72)

nok_ok = add_audit_entry({
    'line': 'STRESS_LINE', 'station': 'SAME_STATION',
    'observation_text': 'Q-Check NOK: Final validation',
    'image_base64': '', 'audit_date': '2026-05-29'
})
T("NOK submit succeeds (no NameError)", nok_ok)

load_audit_entries.clear()
df4 = load_audit_entries()
nok = df4[(df4['line']=='STRESS_LINE') & (df4['observation_text'].str.contains('Q-Check NOK', na=False))]
if not nok.empty:
    nok_img = str(nok['image_base64'].iloc[-1])
    T("NOK has empty image (no leakage)", nok_img in ("", "nan", "None", "NaN"),
      f"actual={repr(nok_img)}")
else:
    T("NOK row found", False)

# ====================================================================
print("\n" + "="*72)
print("PHASE 9: SYNTAX & COMPILATION CHECK")
print("="*72)

import py_compile
for f in ['app.py', 'excel_backend.py', 'setup_environment.py']:
    try:
        py_compile.compile(f, doraise=True)
        T(f"{f} syntax valid", True)
    except py_compile.PyCompileError as e:
        T(f"{f} syntax valid", False, str(e))

# ====================================================================
print("\n" + "="*72)
print("FINAL REPORT")
print("="*72)

passed = sum(1 for _,s,_ in results if s == P)
failed = sum(1 for _,s,_ in results if s == F)
total = len(results)

print(f"\n  Total Tests:    {total}")
print(f"  Passed:         {passed}")
print(f"  Failed:         {failed}")
print(f"  Pass Rate:      {passed/total*100:.1f}%")
print()

if failed == 0:
    print("  *** VERDICT: A) PRODUCTION READY ***")
else:
    print("  *** VERDICT: B) NOT PRODUCTION READY ***")
    print("  Failed tests:")
    for name, status, detail in results:
        if status == F:
            print(f"    X {name}: {detail}")
